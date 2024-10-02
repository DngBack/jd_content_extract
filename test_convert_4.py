from openpyxl import load_workbook

def excel_to_markdown_with_merge(file_path):
    wb = load_workbook(file_path, data_only=True)
    markdown_output = ""

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        markdown_output += f"## {sheet_name}\n\n" 

        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                if cell.value is not None:
                    position = f"{cell.column_letter}{cell.row}"
                    is_merged = any(cell.coordinate in merged_cell for merged_cell in ws.merged_cells.ranges)
                    if is_merged:
                        for merged_cell in ws.merged_cells.ranges:
                            if cell.coordinate in merged_cell:
                                value = ws[merged_cell.start_cell.coordinate].value
                                addresses = [f"{ws.cell(row=r, column=c).coordinate}" 
                                             for r in range(merged_cell.min_row, merged_cell.max_row + 1)
                                             for c in range(merged_cell.min_col, merged_cell.max_col + 1)]
                                address_list = ', '.join(addresses) 
                                row_data.append(f"{value} (Addresses: [{address_list}])")
                                break
                    else:
                        row_data.append(f"{cell.value} ({position})")

            if row_data:
                markdown_output += "| " + " | ".join(row_data) + " |\n"

        markdown_output += "\n"  

    return markdown_output

file_path = '318_1_求人票（総務部）2022.11新規ver2 (2).xlsx'  
markdown_string = excel_to_markdown_with_merge(file_path)
print(markdown_string)

with open('output_4.md', 'w', encoding='utf-8') as f:
    f.write(markdown_string)
