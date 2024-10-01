import openpyxl

def is_homogeneous(cells):
    """Kiểm tra xem các ô có đồng nhất không (tất cả các giá trị giống nhau)."""
    values = [cell for cell in cells if cell is not None]
    return len(set(values)) == 1

def extract_excel_data(sheet):
    """Trả về dữ liệu từ một sheet Excel."""
    merged_ranges = sheet.merged_cells.ranges
    data = []

    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            value = cell.value
            # Nếu ô thuộc một phạm vi được gộp, sử dụng giá trị của ô gốc
            for merged_range in merged_ranges:
                if cell.coordinate in merged_range:
                    # Sửa lỗi bằng cách sử dụng cú pháp đúng để truy cập giá trị của ô
                    value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            row_data.append(value)
        data.append(row_data)

    return data

def find_structural_anchors(data, k):
    """Tìm các hàng và cột không đồng nhất."""
    row_anchors = []
    col_anchors = []

    # Tìm các hàng không đồng nhất
    for i, row in enumerate(data):
        if not is_homogeneous(row):
            row_anchors.append(i)

    # Tìm các cột không đồng nhất
    for j in range(len(data[0])):
        col = [row[j] for row in data]
        if not is_homogeneous(col):
            col_anchors.append(j)

    # Xác định hàng và cột trong phạm vi k đơn vị từ các neo cấu trúc
    selected_rows = set()
    selected_cols = set()

    for r in row_anchors:
        for i in range(max(0, r - k), min(len(data), r + k + 1)):
            selected_rows.add(i)

    for c in col_anchors:
        for j in range(max(0, c - k), min(len(data[0]), c + k + 1)):
            selected_cols.add(j)

    return sorted(list(selected_rows)), sorted(list(selected_cols))

def data_to_markdown(data, selected_rows, selected_cols):
    """Chuyển đổi dữ liệu thành bảng markdown."""
    markdown = "| " + " | ".join([str(data[0][j]) for j in selected_cols]) + " |\n"
    markdown += "| " + " | ".join(["---"] * len(selected_cols)) + " |\n"

    for i in selected_rows:
        row = [str(data[i][j]) if data[i][j] is not None else "" for j in selected_cols]
        markdown += "| " + " | ".join(row) + " |\n"

    return markdown

def extract_all_sheets_to_markdown(file_path, k=1):
    """Chuyển đổi tất cả các sheet trong file Excel thành markdown."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    markdown_result = ""

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Đọc dữ liệu từ sheet
        data = extract_excel_data(sheet)

        # Tìm các hàng và cột không đồng nhất
        selected_rows, selected_cols = find_structural_anchors(data, k)

        # Chuyển đổi thành markdown
        markdown_output = data_to_markdown(data, selected_rows, selected_cols)

        # Thêm tên sheet và markdown của sheet vào kết quả
        markdown_result += f"# {sheet_name}\n\n" + markdown_output + "\n\n"

    return markdown_result

# Ví dụ sử dụng
file_path = "318_1_求人票（総務部）2022.11新規ver2 (2).xlsx"
k_value = 1  # Phạm vi lân cận neo cấu trúc
markdown_result = extract_all_sheets_to_markdown(file_path, k_value)

# In kết quả markdown
print(markdown_result)

# Hoặc lưu vào file markdown
with open("output_2.md", "w", encoding="utf-8") as f:
    f.write(markdown_result)
