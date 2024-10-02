from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def excel_to_markdown_with_merge(file_path):
    wb = load_workbook(file_path, data_only=True)
    markdown_output = ""
    merged_cells_addresses = []  # List to store merged cell addresses

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        markdown_output += f"## {sheet_name}\n\n" 

        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                if cell.value is not None:
                    position = f"{cell.column_letter}{cell.row}"
                    is_merged = any(cell.coordinate in merged_cell for merged_cell in ws.merged_cells.ranges)
                    if is_merged:
                        for merged_cell in ws.merged_cells.ranges:
                            if cell.coordinate in merged_cell:
                                value = ws[merged_cell.start_cell.coordinate].value
                                top_left = merged_cell.start_cell.coordinate
                                bottom_right_row = merged_cell.bounds[3]  
                                bottom_right_col = merged_cell.bounds[2] 
                                bottom_right = f"{get_column_letter(bottom_right_col)}{bottom_right_row}"
                                address_range = f"{top_left}:{bottom_right}"
                                row_data.append(f"{value} (Addresses: [{address_range}])")
                                if address_range not in merged_cells_addresses:
                                    merged_cells_addresses.append(address_range)  # Add address to list
                                break
                    else:
                        row_data.append(f"{cell.value} ({position})")

            if row_data:
                markdown_output += "| " + " | ".join(row_data) + " |\n"

        markdown_output += "\n" 

    return markdown_output, merged_cells_addresses  # Return both markdown output and merged addresses

file_path = '318_1_求人票（総務部）2022.11新規ver2 (2).xlsx' 
markdown_string, merged_addresses = excel_to_markdown_with_merge(file_path)
print(markdown_string)
print("Merged cell addresses:", merged_addresses)

with open('output_6.md', 'w', encoding='utf-8') as f:
    f.write(markdown_string)
