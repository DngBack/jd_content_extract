from __future__ import annotations

import asyncio
import tempfile
from typing import List
from typing import Tuple

import pandas as pd
from domain.parsers.common.models import Cell as CellIR
from domain.parsers.common.models import TableIR
from domain.parsers.common.models import TableMetadata as TableMetadataIR
from domain.parsers.excel.models.excel import Cell
from domain.parsers.excel.models.excel import Header
from domain.parsers.excel.models.excel import SheetFile
from domain.parsers.excel.models.excel import Table
from domain.parsers.excel.models.excel import TableMetadata
from domain.parsers.excel.utils import create_cell_obj
from infra.llm import AzureOpenAIBaseService
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet
from pandas.core.frame import DataFrame
from shared import FileIdentity
from shared.base import BaseModel
from shared.base import BaseService
from shared.logging import get_logger
from shared.settings import Settings

from .header_extractor.processor import HeaderExtractor
from .header_extractor.processor import HeaderExtractorInput
from .summarization.processor import SummarizationExtractor
from .summarization.processor import SummarizationInput
from .table_extractor.processor import TableExtractor
from .table_extractor.processor import TableExtractorInput

logger = get_logger(__name__)


class ParserInput(BaseModel):
    file: FileIdentity


class ParserOutput(BaseModel):
    content: str


class ExcelParser(BaseService):
    llm_model: AzureOpenAIBaseService
    settings: Settings

    @property
    def table_extractor(self) -> TableExtractor:
        """Create table extractor property

        Returns:
            TableExtractor: return TableExtractor object
        """
        return TableExtractor()

    @property
    def summarization_extractor(self) -> SummarizationExtractor:
        """Create table summarization property

        Returns:
            SummarizationExtractor: return SummarizationExtractor object
        """
        return SummarizationExtractor(
            llm_model=self.llm_model,
        )

    @property
    def header_extractor(self) -> HeaderExtractor:
        """Create table header extractor property

        Returns:
            SummarizationExtractor: return HeaderExtractor object
        """
        return HeaderExtractor(
            settings=self.settings,
            llm_model=self.llm_model,
        )

    def process(self, input_data: ParserInput) -> ParserOutput:
        """Parse a PDF document

        Args:
            input_data (ParserInput): input data of parser

        Returns:
            ParserOutput: content of the sheet excel after parsing
        """

        file = input_data.file
        try:
            files = asyncio.run(self.__load_content_file(file))
        except Exception as e:
            logger.exception(f'CRITICAL ERROR: error when loading file, error: {e}')
            raise e

        final_summarization = ''
        for file in files:
            input = TableExtractorInput(file=file)
            tables = self.table_extractor.process(input)
            tablesIR = [self.__convert_to_tableIR(table) for table in tables]
            tablesIR = self.header_extractor.process(HeaderExtractorInput(file=tablesIR))
            tables = [self.__tableIR_to_table(table) for table in tablesIR]
            _, sheet, sheetname = file
            sheet_data = self.__extract_excel_data(sheet, tables, sheetname)

            input = SummarizationInput(file=sheet_data)
            summarization = self.summarization_extractor.process(input)
            final_summarization += f'Sheet {sheetname}: {summarization} \n'
        return ParserOutput(content=final_summarization)

    def __extract_excel_data(self, sheet: Worksheet, tables: List[Table], sheetname: str) -> SheetFile:
        """Extract sheet data

        Args:
            sheet (Worksheet): sheet data read from openpyxl
            tables (List[Table]): List of table data defined in Table pydantic model
            sheetname (str): name of current sheet

        Returns:
            SheetFile: Sheet data defined in SheetFile pydantic model
        """
        paragraph, images = [], []
        for row in sheet.iter_rows():
            for cell in row:
                try:
                    coord = cell.coordinate
                    color = cell.fill.start_color.rgb
                    content = cell.value
                    if not any([
                        coord in CellRange(table.table_metadata.coord)
                        for table in tables
                    ]) and content is not None:
                        cell_obj = create_cell_obj(coord=coord, content=content, color=color)
                        if type(cell.value) is str:
                            paragraph.append(cell_obj)
                        else:
                            images.append(cell_obj)
                except Exception as e:
                    logger.exception(f"""ERROR: error when creating cell obj in
                                     function extract excel data,
                                     error: {e}value: {content}, coord: {coord}, color: {color}""")
                    continue

        return SheetFile(
            sheetname=sheetname,
            tables=tables,
            paragraph=paragraph,
            images=images,
        )

    def __convert_to_tableIR(self, table: Table) -> TableIR:
        """Convert Table format to TableIR format

        Args:
            table (Table): Table object

        Returns:
            TableIR: TableIR object
        """
        cells, merged_cells = [], []

        def __create_cellIR(cell):
            return CellIR(
                coord=cell.coord,
                content=cell.content,
                color=cell.color,
            )
        try:
            for cell_row in table.cells:
                new_cell_rows = []
                for cell in cell_row:
                    new_cell_rows.append(__create_cellIR(cell))
                cells.append(new_cell_rows)
        except Exception as e:
            logger.exception("""ERROR: error when convert
                             table cells to tableIR cells""")
            raise e
        try:
            for merged_cell in table.merged_cells:
                merged_cells.append(__create_cellIR(merged_cell))
        except Exception as e:
            logger.exception("""ERROR: error when convert table
                             merged cells to tableIR merged cells""")
            raise e

        try:
            metadata = TableMetadataIR(coord=table.table_metadata.coord)
        except Exception as e:
            logger.exception("""ERROR: error when convert table
                             metadata to tableIR metadata """)
            raise e

        tableIR = TableIR(
            table_metadata=metadata,
            cells=cells,
            merged_cell=merged_cells,
        )
        return tableIR

    def __tableIR_to_table(self, tableIR: TableIR) -> Table:
        """Convert TableIR format to Table format

        Args:
            table (Table): TableIR object

        Returns:
            TableIR: Table object
        """
        cells, merged_cells = [], []

        def __create_cell(cell):
            return Cell(
                coord=cell.coord,
                content=cell.content,
                color=cell.color,
            )
        try:
            for batch_cell in tableIR.cells:
                new_batch_cell = []
                for cell in batch_cell:
                    new_batch_cell.append(__create_cell(cell))
                cells.append(new_batch_cell)
        except Exception as e:
            logger.exception("""ERROR: error when convert
                             tableIR cells to table cells""")
            raise e

        try:
            for merged_cell in tableIR.merged_cell:
                merged_cells.append(__create_cell(merged_cell))
        except Exception as e:
            logger.exception("""ERROR: error when convert tableIR
                             merged cells to table merged cells""")
            raise e

        try:
            header = tableIR.table_metadata.header
            if header:
                h_header_cells, v_header_cells = [], []
                for h_header_cell in tableIR.table_metadata.header.horizontal:
                    h_header_cells.append(__create_cell(h_header_cell))
                for v_header_cell in tableIR.table_metadata.header.vertical:
                    v_header_cells.append(__create_cell(v_header_cell))
                header = Header(
                    horizontal=h_header_cells,
                    vertical=v_header_cells,
                )
        except Exception as e:
            logger.exception("""ERROR: error when convert tableIR
                             header to table header""")
            raise e

        try:
            metadata = TableMetadata(
                coord=tableIR.table_metadata.coord,
                header=header,
            )
        except Exception as e:
            logger.exception("""ERROR: error when convert tableIR
                             metadata to table metadata """)
            raise e

        table = Table(
            table_metadata=metadata,
            cells=cells,
            merged_cells=merged_cells,
        )
        return table

    async def __load_content_file(self, file: FileIdentity) -> List[Tuple[DataFrame, Worksheet, str]]:
        """Load file content

        Args:
            file (FileIdentity): File data

        Returns:
            List[Tuple[DataFrame, Worksheet, str]]: Tuple of data format for futher processing
        """

        with tempfile.NamedTemporaryFile(
            delete=True,
            suffix=file.file_name,  # pyright: ignore reportPrivateUsage=none
        ) as tmp_file:
            await file.file.seek(0)
            content = await file.file.read()  # pyright: ignore reportPrivateUsage=none
            tmp_file.write(content)  # pyright: ignore reportPrivateUsage=none
            tmp_file.flush()
            wb = load_workbook(tmp_file.name)

            files = []
            for sheetname in wb.sheetnames:
                excel_df = pd.read_excel(tmp_file.name, header=None, sheet_name=sheetname)
                sheet = wb[sheetname]
                files.append((excel_df, sheet, sheetname))
            return files
