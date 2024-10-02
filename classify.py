from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def excel_to_markdown_with_merge_grouped(file_path):
    wb = load_workbook(file_path, data_only=True)
    markdown_output = ""

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        markdown_output += f"## {sheet_name}\n\n"  # Tiêu đề cho mỗi sheet

        # Duyệt qua từng hàng
        for row in ws.iter_rows():
            row_data = []
            merged_ranges = [merged_cell for merged_cell in ws.merged_cells.ranges]
            skip_cells = set()

            for cell in row:
                if cell.coordinate in skip_cells:
                    continue

                if cell.value is not None:
                    is_merged = any(cell.coordinate in merged_cell for merged_cell in merged_ranges)
                    if is_merged:
                        for merged_cell in merged_ranges:
                            if cell.coordinate in merged_cell:
                                value = ws[merged_cell.start_cell.coordinate].value
                                top_left = merged_cell.start_cell.coordinate
                                bottom_right_row = merged_cell.bounds[3]
                                bottom_right_col = merged_cell.bounds[2]
                                bottom_right = f"{get_column_letter(bottom_right_col)}{bottom_right_row}"
                                address_range = f"{top_left}:{bottom_right}"
                                
                                # Thêm giá trị và vùng hợp nhất vào kết quả hàng
                                row_data.append(f"{value} (Addresses: [{address_range}])")

                                # Thêm tất cả các ô nằm trong vùng hợp nhất vào danh sách cần bỏ qua
                                for row_idx in range(merged_cell.bounds[1], merged_cell.bounds[3] + 1):
                                    for col_idx in range(merged_cell.bounds[0], merged_cell.bounds[2] + 1):
                                        skip_cells.add(f"{get_column_letter(col_idx)}{row_idx}")
                                break
                    else:
                        row_data.append(f"{cell.value} ({cell.coordinate})")

            # Ghi lại dữ liệu của hàng chỉ nếu có dữ liệu
            if row_data:
                markdown_output += "| " + " | ".join(row_data) + " |\n"

        markdown_output += "\n"  # Ngắt dòng giữa các sheet

    return markdown_output

# Ví dụ sử dụng
file_path = '318_1_求人票（総務部）2022.11新規ver2 (2).xlsx'  # Thay thế bằng đường dẫn tới file Excel của bạn
markdown_string = excel_to_markdown_with_merge_grouped(file_path)
print(markdown_string)

# Save or print the markdown result
with open('output_4_grouped.md', 'w', encoding='utf-8') as f:
    f.write(markdown_string)
